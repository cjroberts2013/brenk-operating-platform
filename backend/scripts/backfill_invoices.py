"""One-time backfill of historical ServiceChannel invoices from a UI export.

History can't come from the API (`GET /v3/odata/invoices` is permission-
blocked), so it comes from a Provider Automation → Invoices Excel/CSV
export. Full design: docs/architecture/sc-invoice-webhook-sync.md §8.

Run it **after** the webhook is Active — the merge rules resolve any
overlap (webhook-sourced rows always win; backfill only fills gaps).

Safety model (this writes to a real database):
  * **Dry-run by default.** Nothing is written without `--commit`.
  * The detected header mapping is printed every run; review it on a dry
    run before committing.
  * Target DB is explicit: `--database-url`, else `BACKFILL_DATABASE_URL`,
    else the app's configured `DATABASE_URL` (dev). The resolved host is
    printed (password masked) so a prod run is deliberate.
  * Idempotent: re-running the same export changes nothing the second time.

Typical use:

  # 1. dry run against PROD (review the mapping + summary, write nothing):
  python scripts/backfill_invoices.py backfill/input/*.xlsx \\
      --sc-env production --database-url "$PROD_DATABASE_URL"

  # 2. commit:
  python scripts/backfill_invoices.py backfill/input/*.xlsx \\
      --sc-env production --database-url "$PROD_DATABASE_URL" --commit

Merge rules (per row, keyed on (sc_env, invoice_number, coalesce(wo,-1))):
  * no existing row            → INSERT (source='backfill', sc_invoice_id NULL)
  * existing source='webhook'  → webhook wins; only fill columns still NULL
  * existing source='backfill' → overwrite parsed fields (idempotent re-run)

Work-order mirroring is conservative: invoice state is copied onto a
matching `work_orders` row only where that column is still NULL, so a
webhook or a real submit is never clobbered.
"""

from __future__ import annotations

import argparse
import csv
import glob
import re
import sys
from dataclasses import dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

# --------------------------------------------------------------------------- #
# Header mapping — canonical field -> candidate export header names.
# Matching is done on a normalized key (lowercase, alphanumerics only), so
# "Invoice #", "invoice_number" and "INVOICE NUMBER" all collapse together.
# --------------------------------------------------------------------------- #
HEADER_CANDIDATES: dict[str, list[str]] = {
    "invoice_number": ["invoice number", "invoice #", "invoice no", "number"],
    "sc_invoice_id": ["invoice id", "id"],
    "wo_tracking_number": [
        "tracking number",
        "wo tracking number",
        "work order tracking number",
        "tracking #",
        "work order number",
        "wo number",
        "wo #",
        "tracking",
    ],
    "subscriber_id": ["subscriber id"],
    "provider_id": ["provider id"],
    "location_id": ["location id"],
    "status": ["status", "invoice status"],
    "trade": ["trade", "trade name"],
    "category": ["category", "category name"],
    "description": ["description", "invoice text", "resolution"],
    "currency": ["currency", "currency code"],
    "subtotal": ["subtotal", "sub total", "invoice subtotal"],
    "invoice_tax": ["tax", "invoice tax", "sales tax"],
    "invoice_total": ["total", "invoice total", "amount", "total amount"],
    "approval_code": ["approval code", "approval #", "approval"],
    "batch_number": ["batch number", "batch #", "batch"],
    "comments": ["comments", "comment", "notes"],
    "invoice_date": ["invoice date", "date"],
    "posted_date": ["posted date"],
    "approved_date": ["approved date"],
    "paid_date": ["paid date", "payment date"],
    "last_action_date": ["last action date", "last updated", "updated date", "modified date"],
}

REQUIRED_FIELDS = {"invoice_number"}

INT_FIELDS = {"sc_invoice_id", "wo_tracking_number", "subscriber_id", "provider_id", "location_id"}
DEC_FIELDS = {"subtotal", "invoice_tax", "invoice_total"}
DATE_FIELDS = {"invoice_date", "posted_date", "approved_date", "paid_date", "last_action_date"}

# Canonical SC invoice statuses (match invoice_sync._EVENT_STATUS casing).
_STATUS_CANON: dict[str, str] = {
    "open": "Open",
    "approved": "Approved",
    "onhold": "On Hold",
    "hold": "On Hold",
    "reviewed": "Reviewed",
    "inreview": "Reviewed",
    "rejected": "Rejected",
    "reject": "Rejected",
    "void": "Void",
    "voided": "Void",
    "paid": "Paid",
    "disputed": "Disputed",
    "dispute": "Disputed",
}


def _norm_key(name: str) -> str:
    """Normalize a header for matching: lowercase, alphanumerics only."""
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def build_header_map(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    """Map canonical field -> actual header. Returns (mapping, unmapped headers)."""
    norm_to_actual: dict[str, str] = {}
    for h in headers:
        if h is None:
            continue
        norm_to_actual.setdefault(_norm_key(h), h)

    mapping: dict[str, str] = {}
    used: set[str] = set()
    for field_name, candidates in HEADER_CANDIDATES.items():
        for cand in candidates:
            actual = norm_to_actual.get(_norm_key(cand))
            if actual is not None:
                mapping[field_name] = actual
                used.add(_norm_key(actual))
                break

    unmapped = [h for h in headers if h is not None and _norm_key(h) not in used]
    return mapping, unmapped


# --------------------------------------------------------------------------- #
# Value normalization
# --------------------------------------------------------------------------- #
def norm_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def norm_int(v: Any) -> int | None:
    # SC ids are non-negative; treat any non-digit (commas, "WO-" prefixes)
    # purely as separators.
    s = norm_str(v)
    if s is None:
        return None
    digits = re.sub(r"[^0-9]", "", s)
    if not digits:
        return None
    return int(digits)


def norm_decimal(v: Any) -> Decimal | None:
    """Parse a money cell: strips $ , and surrounding spaces; ()=negative."""
    s = norm_str(v)
    if s is None:
        return None
    neg = s.startswith("(") and s.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", s)
    if cleaned in ("", "-", ".", "-."):
        return None
    try:
        d = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -d if neg else d


_DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%m/%d/%Y %H:%M",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %I:%M %p",
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
)


def norm_date(v: Any) -> datetime | None:
    """Parse an export date cell to a tz-aware UTC datetime."""
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=UTC)
    s = norm_str(v)
    if s is None:
        return None
    # ISO 8601 (with optional Z) first.
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=UTC)
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def norm_status(v: Any) -> str | None:
    s = norm_str(v)
    if s is None:
        return None
    return _STATUS_CANON.get(_norm_key(s), s)


def normalize_row(raw: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    """Apply the header map + per-field normalization to one export row."""
    out: dict[str, Any] = {}
    for field_name, header in mapping.items():
        value = raw.get(header)
        if field_name in INT_FIELDS:
            out[field_name] = norm_int(value)
        elif field_name in DEC_FIELDS:
            out[field_name] = norm_decimal(value)
        elif field_name in DATE_FIELDS:
            out[field_name] = norm_date(value)
        elif field_name == "status":
            out[field_name] = norm_status(value)
        else:
            out[field_name] = norm_str(value)
    return out


# --------------------------------------------------------------------------- #
# File readers
# --------------------------------------------------------------------------- #
def read_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (headers, rows-as-dicts) for a .csv or .xlsx file."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
    elif suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        raise ValueError(f"Unsupported file type: {path.name} (need .csv or .xlsx)")

    # Drop fully-empty leading rows, take the first non-empty as the header.
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return [], []
    headers = [("" if h is None else str(h).strip()) for h in rows[0]]
    records: list[dict[str, Any]] = []
    for r in rows[1:]:
        record = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        records.append(record)
    return headers, records


# --------------------------------------------------------------------------- #
# Summary
# --------------------------------------------------------------------------- #
@dataclass
class BackfillSummary:
    read: int = 0
    inserted: int = 0
    merged: int = 0  # existing row had NULLs filled
    overwritten: int = 0  # existing backfill row updated
    skipped_webhook: int = 0  # webhook row, nothing to fill
    wo_linked: int = 0
    failures: list[tuple[int, str]] = field(default_factory=list)  # (row#, reason)

    @property
    def failure_rate(self) -> float:
        return (len(self.failures) / self.read) if self.read else 0.0


# --------------------------------------------------------------------------- #
# DB apply (imported lazily so the pure helpers above stay import-light)
# --------------------------------------------------------------------------- #
_INVOICE_SCALARS = (
    "sc_invoice_id",
    "wo_tracking_number",
    "subscriber_id",
    "provider_id",
    "location_id",
    "status",
    "trade",
    "category",
    "description",
    "currency",
    "subtotal",
    "invoice_tax",
    "invoice_total",
    "approval_code",
    "batch_number",
    "comments",
    "invoice_date",
    "posted_date",
    "approved_date",
    "paid_date",
    "last_action_date",
)


def apply_records(session, records: list[dict[str, Any]], *, sc_env: str) -> BackfillSummary:
    """Upsert parsed invoice records + mirror onto work_orders. Caller
    controls commit/rollback (so dry-run can roll the whole thing back)."""
    from sqlalchemy import select

    from app.models.invoice import Invoice

    summary = BackfillSummary(read=len(records))

    for i, rec in enumerate(records, start=2):  # row 1 is the header
        number = rec.get("invoice_number")
        if not number:
            summary.failures.append((i, "missing invoice_number"))
            continue
        wo = rec.get("wo_tracking_number")

        try:
            stmt = select(Invoice).where(
                Invoice.sc_env == sc_env,
                Invoice.invoice_number == number,
                Invoice.wo_tracking_number.is_(None)
                if wo is None
                else Invoice.wo_tracking_number == wo,
            )
            existing = session.execute(stmt).scalar_one_or_none()

            if existing is None:
                inv = Invoice(sc_env=sc_env, invoice_number=number, source="backfill")
                for col in _INVOICE_SCALARS:
                    if rec.get(col) is not None:
                        setattr(inv, col, rec[col])
                session.add(inv)
                summary.inserted += 1
            elif existing.source == "webhook":
                filled = False
                for col in _INVOICE_SCALARS:
                    if rec.get(col) is not None and getattr(existing, col) is None:
                        setattr(existing, col, rec[col])
                        filled = True
                summary.merged += int(filled)
                summary.skipped_webhook += int(not filled)
            else:  # existing backfill row → overwrite (idempotent re-run)
                for col in _INVOICE_SCALARS:
                    if rec.get(col) is not None:
                        setattr(existing, col, rec[col])
                summary.overwritten += 1

            if wo is not None and _mirror_to_work_order(session, rec, wo):
                summary.wo_linked += 1
        except Exception as exc:
            # One bad row must not abort the whole run.
            summary.failures.append((i, f"{type(exc).__name__}: {exc}"))

    return summary


def _mirror_to_work_order(session, rec: dict[str, Any], wo_tracking: int) -> bool:
    """Copy invoice state onto the matching WO, but only into NULL columns
    (never clobber a webhook or a real submit). Returns True if a WO matched."""
    from sqlalchemy import select

    from app.models.work_order import WorkOrder

    wo = session.execute(
        select(WorkOrder).where(WorkOrder.sc_work_order_id == wo_tracking)
    ).scalar_one_or_none()
    if wo is None:
        return False

    if wo.sc_invoice_number is None and rec.get("invoice_number"):
        wo.sc_invoice_number = rec["invoice_number"]
    if wo.sc_invoice_status is None and rec.get("status"):
        wo.sc_invoice_status = rec["status"]
    if wo.sc_invoice_total is None and rec.get("invoice_total") is not None:
        wo.sc_invoice_total = rec["invoice_total"]
    if wo.sc_invoice_submitted_at is None:
        submitted = rec.get("posted_date") or rec.get("invoice_date")
        if submitted is not None:
            wo.sc_invoice_submitted_at = submitted
    if rec.get("status") == "Paid":
        paid = rec.get("paid_date") or rec.get("last_action_date")
        if paid is not None:
            if wo.sc_paid_at is None:
                wo.sc_paid_at = paid
            if wo.brenk_paid_at is None:
                wo.brenk_paid_at = paid
    return True


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def _mask_url(url: str) -> str:
    return re.sub(r"://([^:/@]+):([^@]+)@", r"://\1:****@", url)


def _resolve_db_url(arg_url: str | None) -> str:
    import os

    if arg_url:
        return arg_url
    env_url = os.environ.get("BACKFILL_DATABASE_URL")
    if env_url:
        return env_url
    from app.core.config import get_settings

    return get_settings().DATABASE_URL


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Backfill SC invoices from a UI export.")
    parser.add_argument("inputs", nargs="+", help="Export file(s) or glob(s) (.csv / .xlsx)")
    parser.add_argument(
        "--sc-env",
        required=True,
        choices=["sandbox", "production"],
        help="Tags every row and selects which work orders to mirror onto.",
    )
    parser.add_argument("--database-url", default=None, help="Override target DB URL.")
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Actually write. Without this the run is a dry-run (rolled back).",
    )
    parser.add_argument(
        "--max-failure-rate",
        type=float,
        default=0.02,
        help="Exit non-zero if more than this fraction of rows fail (default 0.02).",
    )
    parser.add_argument("--failures-csv", default=None, help="Write per-row failures here.")
    args = parser.parse_args(argv)

    paths: list[Path] = []
    for pattern in args.inputs:
        matched = [Path(p) for p in glob.glob(pattern)]
        paths.extend(matched or [Path(pattern)])
    paths = [p for p in paths if p.exists()]
    if not paths:
        print("No input files found.", file=sys.stderr)
        return 2

    # Read + map + normalize every file.
    all_records: list[dict[str, Any]] = []
    for path in paths:
        headers, raw_rows = read_rows(path)
        mapping, unmapped = build_header_map(headers)
        missing_required = REQUIRED_FIELDS - mapping.keys()

        print(f"\n=== {path.name} ===")
        print(f"  rows: {len(raw_rows)}")
        print("  detected mapping:")
        for field_name in HEADER_CANDIDATES:
            actual = mapping.get(field_name)
            print(f"    {field_name:<22} <- {actual if actual else '(none)'}")
        if unmapped:
            print(f"  unmapped columns: {', '.join(unmapped)}")
        if missing_required:
            print(f"  ERROR: required field(s) unmapped: {', '.join(sorted(missing_required))}")
            print("  Fix HEADER_CANDIDATES or the export, then re-run.")
            return 2

        all_records.extend(normalize_row(r, mapping) for r in raw_rows)

    db_url = _resolve_db_url(args.database_url)
    print(f"\nTarget DB: {_mask_url(db_url)}")
    print(f"sc_env: {args.sc_env}   mode: {'COMMIT' if args.commit else 'DRY-RUN (no writes)'}")

    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session

    engine = create_engine(db_url)
    summary: BackfillSummary
    with Session(engine) as session:
        summary = apply_records(session, all_records, sc_env=args.sc_env)
        if args.commit:
            session.commit()
        else:
            session.rollback()
    engine.dispose()

    print("\n--- summary ---")
    print(f"  read:            {summary.read}")
    print(f"  inserted:        {summary.inserted}")
    print(f"  merged (filled): {summary.merged}")
    print(f"  overwritten:     {summary.overwritten}")
    print(f"  skipped(webhook):{summary.skipped_webhook}")
    print(f"  work orders linked: {summary.wo_linked}")
    print(f"  failures:        {len(summary.failures)} ({summary.failure_rate:.1%})")
    if not args.commit:
        print("  (dry-run — nothing was written; re-run with --commit to apply)")

    if summary.failures:
        out = args.failures_csv or "backfill_failures.csv"
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["row", "reason"])
            w.writerows(summary.failures)
        print(f"  failure detail: {out}")

    if summary.failure_rate > args.max_failure_rate:
        print(
            f"\nFAIL: failure rate {summary.failure_rate:.1%} exceeds {args.max_failure_rate:.1%}.",
            file=sys.stderr,
        )
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
