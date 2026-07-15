"""Build an Excel (.xlsx) workbook of store locations.

Pure, DB-free: the endpoint gathers the rows (locations + address + active
gate codes + WO metrics) and hands them here as plain dicts keyed by the
column keys below; this module just renders them into a styled sheet. Keeps
the openpyxl details unit-testable without HTTP or a database.

`COLUMNS` is the single source of truth for the sheet's shape — the header
labels, their order, and the dict keys the endpoint must populate.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# (header label, row-dict key). Order defines the column order. Kept to the
# essentials Daryl asked for — store number, name, address. Add columns here
# (and populate the key in the endpoint's row builder) if the export needs
# to grow later.
COLUMNS: list[tuple[str, str]] = [
    ("Store #", "store_id"),
    ("Name", "name"),
    ("Address", "address"),
]

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def build_locations_workbook(rows: list[dict[str, Any]]) -> bytes:
    """Render location row dicts into a styled .xlsx and return its bytes.

    Each row dict is looked up by the keys in COLUMNS; missing keys render
    as an empty cell. Adds a bold frozen header, an auto-filter, and
    reasonable column widths so the file is usable as-is in Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Locations"

    headers = [label for label, _ in COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([_cell(row.get(key)) for _, key in COLUMNS])

    # Freeze the header row and enable filtering/sorting across all columns.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    for idx, (label, key) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _width_for(label, key)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell(value: Any) -> Any:
    """Coerce a value into something Excel renders cleanly.

    None -> "" (empty cell). Dates/strings/numbers pass through; anything
    exotic is stringified so openpyxl never chokes on an unexpected type.
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    # date/datetime render natively; everything else becomes text.
    if hasattr(value, "isoformat"):
        return value
    return str(value)


def _width_for(label: str, key: str) -> int:
    """A sensible fixed column width — wide for free-text, snug for ids."""
    wide = {"name", "address"}
    if key in wide:
        return 44
    return max(12, len(label) + 2)
