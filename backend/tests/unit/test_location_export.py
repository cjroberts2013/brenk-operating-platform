"""Unit tests for the locations Excel workbook builder (pure, no DB)."""

from io import BytesIO

from openpyxl import load_workbook

from app.services.location_export import COLUMNS, build_locations_workbook


def _load(content: bytes):
    return load_workbook(BytesIO(content))


def test_workbook_has_header_and_rows() -> None:
    rows = [
        {
            "store_id": "0751",
            "name": "CUBESMART AUSTIN STASSNEY",
            "address": "4900 East Stassney Lane, Austin, TX 78744",
        }
    ]
    wb = _load(build_locations_workbook(rows))
    ws = wb.active
    assert ws.title == "Locations"

    header = [c.value for c in ws[1]]
    assert header == [label for label, _ in COLUMNS]
    assert header == ["Store #", "Name", "Address"]

    data = [c.value for c in ws[2]]
    assert data == [
        "0751",
        "CUBESMART AUSTIN STASSNEY",
        "4900 East Stassney Lane, Austin, TX 78744",
    ]

    # Header is frozen and an auto-filter is set over all columns.
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_workbook_missing_keys_and_none_render_blank() -> None:
    wb = _load(build_locations_workbook([{"store_id": "X1"}]))
    ws = wb.active
    values = [c.value for c in ws[2]]
    assert len(values) == len(COLUMNS)
    assert values[0] == "X1"
    assert all(v in ("", None) for v in values[1:])


def test_empty_export_is_valid_workbook_with_header_only() -> None:
    wb = _load(build_locations_workbook([]))
    ws = wb.active
    assert ws.max_row == 1  # header only
    assert [c.value for c in ws[1]] == [label for label, _ in COLUMNS]
